from django.core.management.base import BaseCommand, CommandError
import csv
from itertools import islice
import openpyxl
from transactions.models import Client, Transaction
from transactions.serializers import ClientSerializer, TransactionSerializer
from ...models import ClientError, TransactionError, JobExecution
import datetime
from django.db import connection

def read_csv_file(file_path, batch_size=1000):

    """Loads a CSV file schema using a generator, yielding rows in batches.

    Args:
        filename (str): The name of the CSV file to read.
        batch_size (int): The number of rows to include in each batch.

    Yields:
        list: A list of lists, representing a batch of rows from the CSV file.
    """
        
    with open(file_path, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        next(csvreader)
        while True:
            batch = list(islice(csvreader, batch_size))
            if not batch:
                break
            yield batch

def read_excel_file(file_path, batch_size=1000):
    """Loads an Excel file schema using a generator, yielding rows in batches.

    Args:
        filename (str): The name of the Excel file to read.
        batch_size (int): The number of rows to include in each batch.

    Yields:
        list: A list of lists, representing a batch of rows from the Excel file.
    """

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active 

    row_count = worksheet.max_row
    start_row = 2

    while start_row <= row_count:
        end_row = min(start_row + batch_size - 1, row_count)
        batch = worksheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True)
        yield batch
        start_row = end_row + 1

def load_client_data(job_id, source_path):
    """Loads and validates client data into the target table in batches.

    Args:
        source_path (str): The path to the folder containing the client csv file.
        job_id (str): The id of the etl job.

    Yields:
        list: A tuple containing the number of valid and invalid client records.
    """
    total_valid = 0
    total_invalid = 0
    Client.objects.all().delete()
    for batch in read_csv_file(source_path+"\\clients.csv"):
        valid_data = []
        invalid_data = []
        for client_id, name, email, date_of_birth, country, account_balance in batch:
            # Assuming the row contains data in the correct format
            data = {
                "client_id": client_id,
                "name": name,
                "email": email,
                "date_of_birth": date_of_birth,
                "country": country,
                "account_balance": account_balance
            }

            serializer = ClientSerializer(data=data)
            if serializer.is_valid():
                valid_data.append(Client(
                    client_id=client_id,
                    name=name,
                    email=email,
                    date_of_birth=date_of_birth,
                    country= country,
                    account_balance=account_balance
                ))
            else:
                invalid_data.append(ClientError(
                    job_id= f"{job_id}_Client",
                    client_id=client_id,
                    name=name,
                    email=email,
                    date_of_birth=date_of_birth,
                    country= country,
                    account_balance=account_balance,
                    error = str(serializer.errors)
                ))
        Client.objects.bulk_create(valid_data)
        total_valid = total_valid + len(valid_data)
        ClientError.objects.bulk_create(invalid_data)
        total_invalid = total_invalid + len(invalid_data)
    return(total_valid, total_invalid)


def load_transaction_data(job_id, source_path):
    """Loads and validates transaction data into the target table in batches.

    Args:
        source_path (str): The path to the folder containing the transactions excel file.
        job_id (str): The id of the etl job.

    Yields:
        list: A tuple containing the number of valid and invalid transaction records.
    """
    total_valid = 0
    total_invalid = 0
    Transaction.objects.all().delete()
    for batch in read_excel_file(source_path+"\\transactions.xlsx"):
        valid_data = []
        invalid_data = []
        for transaction_id, client_id, transaction_type, transaction_date, amount, currency in batch:
            # Assuming the row contains data in the correct format
            data = {
                "transaction_id": transaction_id,
                "client_id": client_id,
                "transaction_type": transaction_type,
                "transaction_date": transaction_date,
                "amount": amount,
                "currency": currency
            }

            serializer = TransactionSerializer(data=data)
            if serializer.is_valid():
                valid_data.append(Transaction(
                    transaction_id=transaction_id,
                    client_id=client_id,
                    transaction_type=transaction_type,
                    transaction_date=transaction_date,
                    amount= amount,
                    currency=currency
                ))
            else:
                invalid_data.append(TransactionError(
                    job_id= f"{job_id}_Transaction",
                    transaction_id=transaction_id,
                    client_id=client_id,
                    transaction_type=transaction_type,
                    transaction_date=transaction_date,
                    amount= amount,
                    currency=currency,
                    error = str(serializer.errors)
                ))
        Transaction.objects.bulk_create(valid_data)
        total_valid = total_valid + len(valid_data)
        TransactionError.objects.bulk_create(invalid_data)
        total_invalid = total_invalid + len(invalid_data)
    return(total_valid, total_invalid)

class Command(BaseCommand):
    help = "Loads data into the database in a full load manner"

    def add_arguments(self, parser):
        # indicates the path to the folder containing the clients and transactions data.
        parser.add_argument("--source-path", type=str)

    def handle(self, *args, **options):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        job_id = f"ETL_{timestamp}"
        source_path = options.get("source_path")
        if source_path is None:
           raise CommandError("Argument --source-path is mandatory")
        print("ETL Job: ", job_id)
        print("-----------------------------------------------")
        # Loading Client CSV File
        print("Loading Client Data....")
        start_time = datetime.datetime.now()
        client_valid, client_invalid = load_client_data(job_id, source_path)
        end_time = datetime.datetime.now()
        elapsed_time = end_time - start_time
        print(f"Loaded {client_valid} clients successfully. {client_invalid} clients were not valid. Time Elapsed: {elapsed_time}.")
        JobExecution(
            job_id = f"{job_id}_Client",
            job_name = "Client",
            successful_records = client_valid,
            failed_records = client_invalid
        ).save()
        # Loading Transaction Excel File
        print("Loading Transaction Data....")
        start_time = datetime.datetime.now()
        transaction_valid, transaction_invalid = load_transaction_data(job_id, source_path)
        end_time = datetime.datetime.now()
        elapsed_time = end_time - start_time
        print(f"Loaded {transaction_valid} transactions successfully. {transaction_invalid} transactions were not valid. Time Elapsed: {elapsed_time}.")
        JobExecution(
            job_id = f"{job_id}_Transaction",
            job_name = "Transaction",
            successful_records = transaction_valid,
            failed_records = transaction_invalid
        ).save()
        # Refreshing Materialized View
        with connection.cursor() as cursor:
            cursor.execute("REFRESH MATERIALIZED VIEW transactions_mv")
