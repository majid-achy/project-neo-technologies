import random
import datetime
import openpyxl

# Function to generate random transaction data
def generate_random_transaction(transaction_id, client_id):
    transaction_types = ["buy", "sell"]
    currencies = ["USD", "EUR", "GBP", "JPY"]

    transaction_type = random.choice(transaction_types)
    transaction_date = datetime.date.today() - datetime.timedelta(days=random.randint(0, 365))
    transaction_date = transaction_date.strftime("%Y-%m-%d")
    amount = round(random.uniform(100, 10000), 2)
    if transaction_type == "sell":
        amount = -amount
    currency = random.choice(currencies)

    return [transaction_id, client_id, transaction_type, transaction_date, amount, currency]

# Generate transaction data based on the number of clients
def generate_transactions(num_clients):
    transactions = []
    transaction_id = 1
    for client_id in range(1, num_clients + 1):
        for _ in range(random.randint(1, 10)):
            transactions.append(generate_random_transaction(transaction_id, client_id))
            transaction_id = transaction_id + 1
    return transactions

# Create an Excel workbook and worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Transactions"

# Write the header row
header = ["transaction_id", "client_id", "transaction_type", "transaction_date", "amount", "currency"]
for col, header_value in enumerate(header, 1):
    sheet.cell(row=1, column=col, value=header_value)


num_clients = 1000  # Adjust the number of clients as needed
transactions = generate_transactions(num_clients)
for row, transaction in enumerate(transactions, 2):
    for col, value in enumerate(transaction, 1):
        sheet.cell(row=row, column=col, value=value)


workbook.save("transactions.xlsx")